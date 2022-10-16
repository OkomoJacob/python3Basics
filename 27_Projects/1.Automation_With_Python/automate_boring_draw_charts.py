import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']


# A for loop to iterate through the sheet rows from top to down
for row in range(2, sheet.max_row + 1): # + 1 since range ruturns till the last number minus 1 e.g range(1, 3) returns 1, 2
    # print(row)
    cell = sheet.cell(row, 3)
    # multiply price by 90% to reduce it by 10%
    corrected_price = cell.value * 0.9
    # Append a new column with the corrected_price column
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Reference to select a range of values to be plotted in our bar chart
valus = Reference(sheet,    
        min_row = 2, 
        max_row = sheet.max_row,
        min_col=4,
        max_col=4)
# Create an instance of a Bar chart
chart = BarChart()
chart.add_data(valus)
# Add the chart to our sheet
sheet.add_chart(chart, 'e2')


wb.save('transactions.xlsx')

