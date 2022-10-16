import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

'''
Sheet Methods
=============
1.To print the id of the attribute
cell = sheet.cell(1, 1)
print(cell.value)
cell = sheet['a1']
# Or similarly
2.To get the number of rows in te sheet
print(sheet.max_row)
'''

# Accessing a prticular cell from the numerous cells in the excel sheet
cell = sheet['a1']
cell = sheet.cell(2, 3)
print(cell.value)

# A for loop to iterate through the sheet rows from top to down
for row in range(2, sheet.max_row + 1): # + 1 since range ruturns till the last number minus 1 e.g range(1, 3) returns 1, 2
    # print(row)
    cell = sheet.cell(row, 3)
    # multiply price by 90% to reduce it by 10%
    corrected_price = cell.value * 0.9
    # Append a new column with the corrected_price column
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('transactions2.xlsx')
